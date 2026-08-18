import customtkinter as ctk
from database import get_total_revenue, get_total_sales, get_product_count, get_low_stock_count, get_most_sold_products

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter

from tkinter import filedialog, messagebox
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import (
    get_total_revenue,
    get_total_sales,
    get_product_count,
    get_most_sold_products,
    get_low_stock_count
)


class ReportsPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)

        label = ctk.CTkLabel(
            self,
            text="Reports",
            font=("Inter", 28, "bold")
        )
        label.pack(pady=5)

        # =========================
        # Top Frame
        # =========================

        self.top_frame = ctk.CTkFrame(
            self,
            height=130
        )
        self.top_frame.pack(
            fill="x",
            padx=20,
            pady=10
        )

        self.top_frame.pack_propagate(False)

        # Statistics cards

        cards = [
            ("Revenue", "revenue"),
            ("Total Sales", "sales"),
            ("Products", "products"),
            ("Low Stock", "low_stock")
        ]

        self.card_values = {}
        for title, key in cards:

            card = ctk.CTkFrame(
                self.top_frame,
                height=90,
                width=220
            )

            card.pack(
                side="left",
                expand=True,
                padx=10,
                pady=15
            )

            card.pack_propagate(False)

            title_label = ctk.CTkLabel(
                card,
                text=title,
                font=("Arial", 16, "bold")
            )
            title_label.pack(
                pady=(10, 0)
            )

            value_label = ctk.CTkLabel(
                card,
                text="0",
                font=("Arial", 24, "bold")
            )
            value_label.pack(
                pady=5
            )
            self.card_values[key] = value_label
        # =========================
        # Middle Frame
        # =========================

        self.middle_frame = ctk.CTkFrame(
            self
        )
        self.middle_frame.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=10
        )

        self.create_best_selling_graph()

        # =========================
        # Bottom Frame
        # =========================

        self.bottom_frame = ctk.CTkFrame(
            self,
            height=110,
            corner_radius=15
        )

        self.bottom_frame.pack(
            fill="x",
            padx=20,
            pady=10
        )

        self.bottom_frame.pack_propagate(False)

        self.load_reports()

        # Export section
        export_card = ctk.CTkFrame(
            self.bottom_frame,
            fg_color="#242424",
            corner_radius=18
        )

        export_card.pack(
            fill="both",
            expand=True,
            padx=10,
            pady=10
        )

        # Left side text
        export_info = ctk.CTkFrame(
            export_card,
            fg_color="transparent"
        )

        export_info.pack(
            side="left",
            padx=25,
            pady=15
        )

        export_title = ctk.CTkLabel(
            export_info,
            text="📄 Generate Report",
            font=("Arial", 16, "bold"),
            text_color="white"
        )

        export_title.pack(
            anchor="w"
        )

        export_subtitle = ctk.CTkLabel(
            export_info,
            text="Export your business analytics as a professional PDF/Excel",
            font=("Arial", 12),
            text_color="#9CA3AF"
        )

        export_subtitle.pack(
            anchor="w",
            pady=(3, 0)
        )

        # Right side buttons
        buttons_frame = ctk.CTkFrame(
            export_card,
            fg_color="transparent"
        )

        buttons_frame.pack(
            side="right",
            padx=25
        )

        # PDF button
        self.export_pdf_btn = ctk.CTkButton(
            buttons_frame,
            text="📄  Export PDF",
            width=150,
            height=42,
            corner_radius=14,
            fg_color="#2563EB",
            hover_color="#1D4ED8",
            font=("Arial", 13, "bold"),
            command=self.export_pdf
        )

        self.export_pdf_btn.pack(
            side="left",
            padx=8
        )

        # Excel button (ready for later)
        self.export_excel_btn = ctk.CTkButton(
            buttons_frame,
            text="📊  Export Excel",
            width=150,
            height=42,
            corner_radius=14,
            fg_color="#16A34A",
            hover_color="#15803D",
            font=("Arial", 13, "bold"),
            command=self.export_excel
        )

        self.export_excel_btn.pack(
            side="left",
            padx=8
        )

    def load_reports(self):

        revenue = get_total_revenue()
        sales = get_total_sales()
        products = get_product_count()
        low_stock = get_low_stock_count()

        self.card_values["revenue"].configure(
            text=f"{revenue:.2f} DH"
        )

        self.card_values["sales"].configure(
            text=str(sales)
        )

        self.card_values["products"].configure(
            text=str(products)
        )

        self.card_values["low_stock"].configure(
            text=str(low_stock)
        )

    def create_best_selling_graph(self):

        data = get_most_sold_products()

        if not data:
            return

        # Data
        products = [item[0] for item in data]
        quantities = [item[1] for item in data]

        max_quantity = max(quantities)

        # Main card
        scroll_frame = ctk.CTkScrollableFrame(
            self.middle_frame,
            fg_color="#242424",
            corner_radius=20
        )

        scroll_frame.pack(
            fill="both",
            expand=True,
            padx=15,
            pady=15
        )

        graph_card = ctk.CTkFrame(
            scroll_frame,
            fg_color="transparent",
            corner_radius=20
        )

        graph_card.pack(
            fill="both",
            expand=True,
            padx=5,
            pady=5
        )

        # Header
        header_frame = ctk.CTkFrame(
            graph_card,
            fg_color="transparent"
        )

        header_frame.pack(
            fill="x",
            padx=25,
            pady=(20, 5)
        )

        title = ctk.CTkLabel(
            header_frame,
            text="🏆  Top Selling Products",
            font=("Arial", 18, "bold"),
            text_color="white"
        )

        title.pack(
            anchor="w"
        )

        subtitle = ctk.CTkLabel(
            header_frame,
            text="Based on quantity sold",
            font=("Arial", 12),
            text_color="#9CA3AF"
        )

        subtitle.pack(
            anchor="w",
            pady=(3, 0)
        )

        # Products container
        products_frame = ctk.CTkFrame(
            graph_card,
            fg_color="transparent"
        )

        products_frame.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=15
        )

        colors = [
            "#3B82F6",
            "#2563EB",
            "#60A5FA",
            "#1D4ED8",
            "#38BDF8"
        ]

        for index, (product, quantity) in enumerate(zip(products, quantities)):

            row = ctk.CTkFrame(
                products_frame,
                fg_color="transparent"
            )

            row.pack(
                fill="x",
                pady=8
            )

            # Product name + quantity
            info_frame = ctk.CTkFrame(
                row,
                fg_color="transparent"
            )

            info_frame.pack(
                fill="x"
            )

            name = ctk.CTkLabel(
                info_frame,
                text=("👑 " if index == 0 else "") + product,
                font=("Arial", 13, "bold"),
                text_color="white"
            )

            name.pack(
                side="left"
            )

            value = ctk.CTkLabel(
                info_frame,
                text=f"{quantity} sold",
                font=("Arial", 12, "bold"),
                fg_color="#1F2937",
                corner_radius=10,
                text_color="#D1D5DB",
                padx=10
            )

            value.pack(
                side="right"
            )

            # Bar background
            bar_bg = ctk.CTkFrame(
                row,
                height=14,
                fg_color="#3A3A3A",
                corner_radius=20
            )

            bar_bg.pack(
                fill="x",
                pady=(8, 0)
            )

            # Filled bar
            fill_width = int(
                (quantity / max_quantity) * 500
            )

            bar_fill = ctk.CTkFrame(
                bar_bg,
                height=14,
                width=fill_width,
                fg_color=colors[index % len(colors)],
                corner_radius=20
            )

            bar_fill.place(
                x=0,
                y=0
            )

    def export_pdf(self):

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[
                ("PDF Files", "*.pdf")
            ]
        )

        if not file_path:
            return

        doc = SimpleDocTemplate(
            file_path
        )

        styles = getSampleStyleSheet()

        elements = []

        # Title
        elements.append(
            Paragraph(
                "Shop Inventory Report",
                styles["Title"]
            )
        )

        elements.append(
            Spacer(1, 20)
        )

        elements.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                styles["Normal"]
            )
        )

        elements.append(
            Spacer(1, 20)
        )

        # Summary
        elements.append(
            Paragraph(
                "Business Summary",
                styles["Heading2"]
            )
        )

        summary_data = [
            ["Metric", "Value"],

            [
                "Total Revenue",
                f"{get_total_revenue()} DH"
            ],

            [
                "Total Sales",
                str(get_total_sales())
            ],

            [
                "Products Count",
                str(get_product_count())
            ],

            [
                "Low Stock Products",
                str(get_low_stock_count())
            ]
        ]

        summary_table = Table(summary_data)

        summary_table.setStyle(
            TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, None),
                ("ALIGN", (0, 0), (-1, -1), "CENTER")
            ])
        )

        elements.append(summary_table)

        elements.append(
            Spacer(1, 20)
        )

        # Top Selling Products
        elements.append(
            Paragraph(
                "Top Selling Products",
                styles["Heading2"]
            )
        )

        top_data = [
            ["Product", "Quantity Sold"]
        ]

        for product, quantity in get_most_sold_products():

            top_data.append(
                [
                    product,
                    str(quantity)
                ]
            )

        top_table = Table(top_data)

        top_table.setStyle(
            TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, None),
                ("ALIGN", (0, 0), (-1, -1), "CENTER")
            ])
        )

        elements.append(top_table)

        doc.build(elements)

        messagebox.showinfo(
            "Export Complete",
            "PDF report created successfully!"
        )

    def export_excel(self):

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel Files", "*.xlsx")
            ]
        )

        if not file_path:
            return

        workbook = Workbook()

        # Summary Sheet
        summary_sheet = workbook.active

        summary_sheet.title = "Summary"

        summary_sheet.append(
            [
                "Metric",
                "Value"
            ]
        )

        summary_sheet.append(
            [
                "Total Revenue",
                get_total_revenue()
            ]
        )

        summary_sheet.append(
            [
                "Total Sales",
                get_total_sales()
            ]
        )

        summary_sheet.append(
            [
                "Products Count",
                get_product_count()
            ]
        )

        summary_sheet.append(
            [
                "Low Stock Products",
                get_low_stock_count()
            ]
        )

        # Top Sellers Sheet
        top_sheet = workbook.create_sheet(
            "Top Sellers"
        )

        top_sheet.append(
            [
                "Product",
                "Quantity Sold"
            ]
        )

        for product, quantity in get_most_sold_products():

            top_sheet.append(
                [
                    product,
                    quantity
                ]
            )

        # Styling
        for sheet in workbook:

            for cell in sheet[1]:

                cell.font = Font(
                    bold=True
                )

                cell.alignment = Alignment(
                    horizontal="center"
                )

            for column in sheet.columns:

                max_length = 0

                letter = column[0].column_letter

                for cell in column:

                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                sheet.column_dimensions[
                    letter
                ].width = max_length + 5

        workbook.save(
            file_path
        )

        messagebox.showinfo(
            "Export Complete",
            "Excel report created successfully!"
        )
